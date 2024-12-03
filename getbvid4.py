import requests
from openpyxl import Workbook
import json
import argparse

# 创建命令行参数解析器
parser = argparse.ArgumentParser(description='Fetch Bilibili video data and save to Excel.')
parser.add_argument('--mid', type=str, default='37737161', help='Bilibili user ID')
parser.add_argument('--season_id', type=str, default='4255440', help='Season ID')
parser.add_argument('--page_num', type=str, default='1', help='Page number')
parser.add_argument('--page_size', type=str, default='100', help='Page size')
parser.add_argument('--output_file', type=str, default='video_data.xlsx', help='Output Excel file name')
args = parser.parse_args()

# 创建一个新的Excel文件
workbook = Workbook()
sheet = workbook.active

# 写入标题行
sheet.cell(row=1, column=1, value="Video Title")
sheet.cell(row=1, column=2, value="BV Number")
sheet.cell(row=1, column=3, value="Play Count")

# 设置Bilibili API的headers
headers = {
    'User-Agent': 'Mozilla/5.0'
}

# 发送请求获取视频列表
params = {
    "mid": args.mid,
    "sort_reverse": "false",
    "season_id": args.season_id,
    "page_num": args.page_num,
    "page_size": args.page_size
}

try:
    # 使用requests库发送请求，设置超时时间为5秒（可调整）
    response = requests.get("https://api.bilibili.com/x/polymer/space/seasons_archives_list", headers=headers,
                            params=params, timeout=5)
    # 检查响应状态码
    if response.status_code == 200:
        try:
            # 解析JSON数据
            data_json = response.json()
            # 检查'code'键是否存在于data_json字典中，并且是否为0（成功）
            if data_json.get('code') == 0 and 'data' in data_json:
                aids = data_json['data'].get('aids', [])
                archives = data_json['data'].get('archives', [])
                # 创建字典方便根据aid查找对应的archive信息（因为原数据既有aids列表又有archives列表结构）
                aid_archive_dict = {archive['aid']: archive for archive in archives}
                for aid in aids:
                    try:
                        archive = aid_archive_dict.get(aid)
                        if archive:
                            video_title = archive['title']
                            bv = archive['bvid']
                            play_count = archive['stat'].get('view', 0)  # 更严谨获取播放量，若不存在设为0
                            # 使用append方法写入Excel，提升效率
                            sheet.append([video_title, bv, play_count])
                        else:
                            print(f"Archive not found for aid: {aid}")
                    except KeyError as e:
                        print(f"Key error occurred: {e}, skipping this video record.")
                    except Exception as e:
                        print(f"An error occurred: {e}")
            else:
                print("Error:", data_json.get('message', 'Unknown error'))
        except json.JSONDecodeError:
            print("Failed to decode JSON data from the response.")
    else:
        print(f"Failed to retrieve data, status code: {response.status_code}")
except requests.RequestException as e:
    print(f"Request error: {e}")

# 保存Excel文件
workbook.save(args.output_file)